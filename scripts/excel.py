import openpyxl
import openpyxl.utils.cell


class Excel:
    def __init__(self, file_path: str, active_sheet: int = 1):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

        wb_sheet = self.workbook.sheetnames

        self.active_sheet = self.workbook[wb_sheet[active_sheet - 1]]

    def check_range(range: any):
        if type(range) not in (str, list):
            raise TypeError("Range must be a type of string or list")

        elif type(range) == str:
            if range.isalpha() or range.isnumeric():
                raise TypeError(
                    "Range string must be a combination of character and number"
                )

        elif type(range) == list:
            if len(range) == 2:
                for i in range:
                    if type(i) not in (str, int):
                        raise TypeError(
                            "Range list can only have a type of string and integer for its values"
                        )

            else:
                raise TypeError("Range list can only have 2 values")

    def convert_range(range: any):
        Excel.check_range(range)

        if type(range) == str:
            column = Excel.check_and_convert_string_value(
                "".join(x for x in range if not x.isdigit())
            )
            row = int("".join(x for x in range if x.isdigit()))

        elif type(range) == list:
            if type(range[0]) == str:
                column = Excel.check_and_convert_string_value(range[0])

            elif type(range[0]) == int:
                column = range[0]

            if type(range[1]) == str:
                row = Excel.check_and_convert_string_value(range[1])

            elif type(range[1]) == int:
                row = range[1]

        return column, row

    def check_and_convert_string_value(value: any):
        if type(value) == str:
            value = [ord(x) - 96 for x in value.lower()]

            new_value = 0
            for i in range(len(value)):
                new_value += value[i] * 26 ** (len(value) - (i + 1))

        return new_value

    def attributes_string(list_of_attributes: any):
        attributes_string = ""
        for i, attribute in enumerate(list_of_attributes):
            if i == 0:
                attributes_string += attribute

            else:
                attributes_string += f", {attribute}"

        return attributes_string

    def get_value_multiple_2d(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        value_array = []
        for row in range(start_row, end_row + 1):
            temp_value_array = []
            for column in range(start_column, end_column + 1):
                temp_value = self.active_sheet.cell(row=row, column=column).value
                temp_value_array.append(temp_value)

            value_array.append(temp_value_array)

        return value_array
